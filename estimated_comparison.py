import openpyxl
import os
import datetime
import pyodbc

path = fr'\\172.16.0.190\Operation\ReservoirDocuments\1_Daily Report from email'

def main():
    
    startDateStr = '2020-09-01'
    endDateStr = '2020-10-01'

    startDate = datetime.datetime.strptime(startDateStr, '%Y-%m-%d')
    endDate = datetime.datetime.strptime(endDateStr, '%Y-%m-%d')

    excelData = dict()

    # while startDate != endDate:
    #     print(startDate)
    #     startDate += datetime.timedelta(days=1)


    conn = pyodbc.connect('Driver={SQL Server};'
    'Server=QPCAVODB2K12;'
    'Database=AVOCET_PRODUCTION;'
    'Trusted_Connection=yes;'
    'UID=sa;'
    'PWD=A@qpc2017;')

    cursor = conn.cursor()
    cursor.execute(f'''SELECT D.[DATETIME], C.WELL, C.ITEM_NAME, C.START_DATETIME, C.END_DATETIME, C.[STATUS], C.[TYPE], ISNULL(T.PROD_OIL_VOL, 0) AS PROD_OIL_VOL, ISNULL(T.PROD_LIQ_VOL, 0) AS PROD_LIQ_VOL, ISNULL(DW.DURATION, 0) AS DURATION
    ,TEST.START_DATETIME, TEST.ITEM_NAME, TEST.LIQ_VOL, TEST.BSW_VOL_FRAC * 100 AS BSW, TEST.OIL_VOL
    FROM
    AVOCET_PRODUCTION.dbo.DATE_INFO AS D
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VI_COMPLETION_ALL_en_US AS C
    ON C.START_DATETIME <= D.[DATETIME]
    AND
    C.END_DATETIME > D.[DATETIME]
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VT_TOTALS_DAY_en_US AS T
    ON
    T.ITEM_ID = C.ITEM_ID
    AND
    T.START_DATETIME = D.[DATETIME]
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VT_DOWNTIME_en_US AS DW
    ON
    DW.ITEM_ID = C.ITEM_ID
    AND
    CAST(DW.START_DATETIME AS DATE) = CAST(D.[DATETIME] AS DATE)
    LEFT JOIN
    VT_WELL_TEST_en_US AS TEST
    ON
    TEST.ITEM_ID = C.ITEM_ID
    AND
    TEST.START_DATETIME = (SELECT TOP 1 START_DATETIME FROM VT_WELL_TEST_en_US WHERE START_DATETIME <= D.[DATETIME] AND ITEM_ID = C.ITEM_ID AND VALID_TEST = 'True' ORDER BY START_DATETIME DESC)
    AND
    TEST.VALID_TEST = 'True'


    WHERE
    D.[DATETIME] >= '{startDateStr}'
    AND
    D.[DATETIME] < '{endDateStr}'
    ORDER BY
    D.[DATETIME]''')
    avocetData = dict()
    for row in cursor:
        if row.DATETIME not in avocetData:
            avocetData[row.DATETIME] = dict()
            excelData[row.DATETIME] = getExcelForDay(row.DATETIME + datetime.timedelta(days=1))
        if row.WELL in avocetData[row.DATETIME]:
            if avocetData[row.DATETIME][row.WELL]["rate"] > row.PROD_OIL_VOL:
                continue
        avocetData[row.DATETIME][row.WELL] = {
            "status": row.STATUS,
            "hrs_online": 24 - (row.DURATION / 60 / 60),
            "rate": int(row.PROD_OIL_VOL),
            "bopd": row.OIL_VOL,
            "bsw": row.BSW,
            "bfpd": row.LIQ_VOL,
            "name": row.ITEM_NAME
        }
            
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Production Date",
     "Excel Name", "Avocet Name",
      "Excel Estimated Oil", "Avocet Estimated Oil",
    "Excel Hrs Online", "Avocet Hrs Online",
    "Excel Test BF", "Avocet Test BF",
    "Excel Test BSW", "Avocet Test BSW",
    "Excel Test BO", "Avocet Test BO",
    "Excel Status", "Avocet Status"])

    for date in excelData:
        for well in excelData[date]:
            if abs(int(avocetData[date][well]["rate"]) - int(excelData[date][well]["rate"])) > 2:
                print(date)
                print(well)
                print(excelData[date][well])
                print(avocetData[date][well])
                ws.append([date.strftime('%Y-%m-%d'),
                 well, avocetData[date][well]["name"],
                 excelData[date][well]["rate"], avocetData[date][well]["rate"],
                 excelData[date][well]["hrs_online"], avocetData[date][well]["hrs_online"],
                 excelData[date][well]["bfpd"], avocetData[date][well]["bfpd"],
                 excelData[date][well]["bsw"], avocetData[date][well]["bsw"],
                 excelData[date][well]["bopd"], avocetData[date][well]["bopd"],
                 excelData[date][well]["status"], avocetData[date][well]["status"]
                 ])
    wb.save(f"Comparison from {startDateStr} to {endDateStr}.xlsx")                




    # for row in cursor:
    #     if row.DATETIME not in excelData:
    #         excelData[row.DATETIME] = getExcelForDay(row.DATETIME + datetime.timedelta(days=1))

    #     if row.WELL in excelData[row.DATETIME] and row.TYPE == 'PRODUCTION' and row.STATUS == 'PRODUCING':
    #         if abs(int(row.PROD_OIL_VOL) - int(excelData[row.DATETIME][row.WELL]["bopd"] * excelData[row.DATETIME][row.WELL]["hrs_online"] / 24)) < 3:
    #             del excelData[row.DATETIME][row.WELL]
    #             continue
    #     elif row.WELL not in excelData[row.DATETIME] and row.TYPE == 'PRODUCTION' and row.STATUS == 'PRODUCING':
    #         print(row.WELL + " " + row.ITEM_NAME + " not found in Excel")
            
    #     elif row.WELL in excelData[row.DATETIME] and row.TYPE == 'PRODUCTION':
    #         if excelData[row.DATETIME][row.WELL]["hrs_online"] > 0: 
    #             if abs(int(row.PROD_OIL_VOL) - int(excelData[row.DATETIME][row.WELL]["bopd"] * excelData[row.DATETIME][row.WELL]["hrs_online"] / 24)) > 3:
    #                 print(row.PROD_OIL_VOL)
    #                 print(row.WELL)
    #                 print(excelData[row.DATETIME][row.WELL])
    


def getExcelForDay(date):

    excelData = dict()
    currentDay = fr'\{date.year}\{date.strftime("%m")}-{date.strftime("%b")}\{date.strftime("%d")}'

    for file in os.listdir(path + currentDay):
        if file.endswith(".xlsm") and not file.startswith("~$"):
            print(os.path.join(path + currentDay, file))
            wb = openpyxl.load_workbook(os.path.join(path + currentDay, file), data_only=True, read_only=True)
            if "WELL DATA" in wb.sheetnames:
                ws = wb["WELL DATA"]
                for row in ws.iter_rows(min_row=6, min_col=2, values_only=True):
                    if row[0]:
                        excelData[row[0]] = {
                            "status": row[2],
                            "hrs_online": row[3],
                            "bopd": row[37],
                            "bsw": row[35],
                            "bfpd": row[34],
                            "rate": row[3] / 24 * row[37]
                        }
    return excelData


    


                





main()